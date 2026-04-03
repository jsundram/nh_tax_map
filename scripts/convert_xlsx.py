# /// script
# requires-python = ">=3.10"
# dependencies = ["openpyxl"]
# ///
"""Convert NH DRA tax rate xlsx files (2022-2025) to JSON for the standalone map."""
import json
import sys
from pathlib import Path

import openpyxl


REWRITES = {
    "Erving's Grant (U)": "Ervings Location",
    "Atkinson & Gilmanton Academy Grant (U)": "Atkinson & Gilmanton",
    "Second College Grant (U)": "Second College",
    "Pinkham's Grant (U)": "Pinkham's Grant",
    "Low & Burbank's Grant (U)": "Low & Burbanks",
    "Thomsom & Meserve's Purchase (U)": "Thompson & Meserve",
    "Thompson & Meserve's Purchase (U)": "Thompson & Meserve",
    "Wentworth's Location": "Wentworths Location",
    "Wentworth Location": "Wentworths Location",
}


def normalize(s):
    if s in REWRITES:
        return REWRITES[s]
    return s.replace(" (U)", "").replace("\u2019", "")


# Column mappings: (Municipality, Date, Valuation, Val+Util, Municipal, County, StateEd, LocalEd, TotalRate, TotalCommitment)
# 2022 has a sparse layout with gaps; 2023-2025 are contiguous A-J.
COLUMN_MAP_2022 = (0, 3, 6, 7, 8, 9, 10, 12, 13, 14)  # A, D, G, H, I, J, K, M, N, O
COLUMN_MAP_DEFAULT = (0, 1, 2, 3, 4, 5, 6, 7, 8, 9)    # A, B, C, D, E, F, G, H, I, J

# Sheet name patterns and header row offsets per year
YEAR_CONFIG = {
    2022: {"sheet": "2022 Municipal Tax Rates", "header_row": 5, "cols": COLUMN_MAP_2022},
    2023: {"sheet": "2023 Municipal Tax Rates", "header_row": 4, "cols": COLUMN_MAP_DEFAULT},
    2024: {"sheet": "2024 Municipal Tax Rates", "header_row": 4, "cols": COLUMN_MAP_DEFAULT},
    2025: {"sheet": "2025 Municipal Tax Rates", "header_row": 4, "cols": COLUMN_MAP_DEFAULT},
}

# Expected xlsx filenames per year
YEAR_FILES = {
    2022: "2022-municpal-and-village-tax-rates.xlsx",
    2023: "2023-municipal-and-village-tax-rates.xlsx",
    2024: "2024-municipal-and-village-tax-rates.xlsx",
    2025: "2025-municipal-and-village-district-tax-rates.xlsx",
}


def convert_year(xlsx_path, year):
    """Convert a single year's xlsx to a list of records."""
    config = YEAR_CONFIG[year]
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    # Find the sheet (try exact match, then partial match)
    sheet_name = config["sheet"]
    if sheet_name not in wb.sheetnames:
        # Try partial match
        for sn in wb.sheetnames:
            if "Municipal Tax Rate" in sn:
                sheet_name = sn
                break
        else:
            wb.close()
            raise ValueError(f"No municipal tax rates sheet found in {xlsx_path}. Sheets: {wb.sheetnames}")

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    cols = config["cols"]
    data_start = config["header_row"] + 1  # data starts right after header (0-indexed)

    records = []
    for row in rows[data_start:]:
        muni = row[cols[0]]
        if not muni or not str(muni).strip():
            continue

        name = normalize(str(muni).strip())

        def num(idx):
            v = row[cols[idx]] if cols[idx] < len(row) else None
            if v is None:
                return 0
            try:
                return float(v)
            except (ValueError, TypeError):
                return 0

        records.append({
            "Municipality": name,
            "Date": str(row[cols[1]]) if row[cols[1]] else None,
            "Valuation": num(2),
            "Valuation+Utilities": num(3),
            "Municipal": num(4),
            "County": num(5),
            "StateEd": num(6),
            "LocalEd": num(7),
            "TotalRate": num(8),
            "TotalCommitment": num(9),
        })

    return records


def main():
    data_dir = Path(__file__).parent.parent / "data"

    if len(sys.argv) > 1:
        # Single file mode: convert_xlsx.py <xlsx> [output.json] [year]
        xlsx = Path(sys.argv[1])
        year = int(sys.argv[3]) if len(sys.argv) > 3 else 2022
        output = Path(sys.argv[2]) if len(sys.argv) > 2 else data_dir / "tax_rates.json"
        records = convert_year(xlsx, year)
        with open(output, "w") as f:
            json.dump(records, f, indent=2, default=str)
        print(f"Wrote {len(records)} records to {output}")
        return

    # Multi-year mode: convert all available years
    all_years = {}
    for year, filename in sorted(YEAR_FILES.items()):
        xlsx_path = data_dir / filename
        if not xlsx_path.exists():
            print(f"  Skipping {year}: {filename} not found")
            continue

        records = convert_year(xlsx_path, year)
        all_years[year] = records
        print(f"  {year}: {len(records)} records from {filename}")

    if not all_years:
        print("No xlsx files found!")
        return

    # Write per-year JSON files
    for year, records in all_years.items():
        outpath = data_dir / f"tax_rates_{year}.json"
        with open(outpath, "w") as f:
            json.dump(records, f, indent=2, default=str)
        print(f"Wrote {outpath.name}")

    # Write the default (latest year) as tax_rates.json
    latest = max(all_years.keys())
    outpath = data_dir / "tax_rates.json"
    with open(outpath, "w") as f:
        json.dump(all_years[latest], f, indent=2, default=str)
    print(f"\nDefault tax_rates.json set to {latest} ({len(all_years[latest])} records)")


if __name__ == "__main__":
    main()
